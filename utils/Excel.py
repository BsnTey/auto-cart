import openpyxl
import json
from threading import Lock


class Excel:
    lock = Lock()

    def __init__(self, path_in, path_out):
        self.path_in = path_in
        self.path_out = path_out

    def check_rows_exist(self):
        with self.lock:
            workbook = openpyxl.load_workbook(self.path_in)
            sheet = workbook.active

            rows = list(sheet.iter_rows())

            if len(rows) == 0:
                workbook.close()
                return False

            row_c1 = rows[0][2].value
            row_c2 = rows[0][3].value

            if (row_c1 is None or row_c1 == 'null') or \
                    (row_c2 is None or row_c2 == 'null'):
                workbook.close()
                return False

            workbook.close()
            return True

    def get_excel(self):
        with self.lock:
            workbook = openpyxl.load_workbook(self.path_in)

            sheet = workbook.active

            row_number = 1
            row_values = []
            for cell in sheet[1]:
                row_values.append(cell.value)

            sheet.delete_rows(row_number)
            workbook.save(self.path_in)

            if "[" in row_values[2]:
                email, password_imap, cookie, access_token, refresh_token, x_user, device_id, install_id = row_values
            else:
                email, password_imap, password, cookie, access_token, refresh_token, x_user, device_id, install_id = row_values

            workbook.close()
            # return cookie
            return [email, password_imap, cookie, access_token, refresh_token, x_user, device_id, install_id]

    def finally_add(self, write_values, path="OUT"):
        if path == "OUT":
            path_out_save = self.path_out
        else:
            path_out_save = self.path_in

        with self.lock:
            workbook = openpyxl.load_workbook(path_out_save)
            sheet = workbook.active

            row = [
                write_values[0],
                write_values[1],
                json.dumps(write_values[2], ensure_ascii=False).replace("\\", "").strip('"'),
                write_values[3],
                write_values[4],
                write_values[5],
                write_values[6],
                write_values[7]
            ]

            sheet.append(row)
            workbook.save(path_out_save)
            workbook.close()
